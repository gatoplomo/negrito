






import pandas as pd

from pathlib import Path

import os


import sys


#from flask import Flask
import json 
   

import openpyxl 
  
from openpyxl.chart import LineChart,Reference 


from datetime import date

from openpyxl import Workbook


from openpyxl.chart import (
    LineChart,
    Reference,
)


from openpyxl.chart.axis import DateAxis



# Setup flask server
#app = Flask(__name__) 
  



a=sys.argv[1]
b=sys.argv[2]
c=sys.argv[3]

csv = pd.read_csv(r'/home/tomas/Escritorio/negrito/Datos2/'+sys.argv[2]+'/'+a+'.csv')

df = pd.DataFrame(csv)
  
# displaying the DataFrame
#display(df)


# create excel writer object
writer = pd.ExcelWriter(r'/home/tomas/Escritorio/negrito/Datos2/'+sys.argv[2]+'/'+a+'.xlsx')
# write dataframe to excel
df.to_excel(writer)
# save the excel
writer.save()
print('DataFrame is written successfully to Excel File.')







cantidad=sys.argv[3];



  
wb = openpyxl.Workbook() 
sheet = wb.active 


 
# Give the location of the file
path = '/home/tomas/Escritorio/negrito/Datos2/'+sys.argv[2]+'/'+a+'.xlsx'
 
# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)
 
# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
 
# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.
 
# Note: The first row or
# column integer is 1, not 0.
 
# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj.cell(row = 5, column = 3)
 
# Print value of cell object
# using the value attribute
print(cell_obj.value)


data=[]
data2=[]

for i in range(int(cantidad)): 
    cell_obj = sheet_obj.cell(row = i+2, column = 3)
    cell_obj2 = sheet_obj.cell(row = i+2, column = 5)
    data.append(cell_obj.value) 
    data2.append(cell_obj2.value)
  

#print(data);






#values = Reference(sheet, min_col = 3, min_row = 5, 
                         #max_col = 3, max_row = 15) 
  
#chart = LineChart() 
  
#chart.add_data(values) 
  
#chart.title = " LINE-CHART "
  
#chart.x_axis.title = " X-AXIS "
  
#chart.y_axis.title = " Y-AXIS "
  
#sheet.add_chart(chart, "E2") 
  
#wb.save("C:/Users/idcla/Documents/GitHub/propal/Datos/nodo1/09-01-2022-14-32-05.xlsx") 




wb = Workbook()
ws = wb.active



 #[date(2015,9, 1), 40]

rows = [
    ['Valor', 'Hora']
]

i=0
for row in data:

    rows.append([row,data2[i]])
    i=i+1;
print(rows)


for row in rows:
    ws.append(row)
    print(row)


chart = LineChart()
chart.title = "Fecha/Central/Sensor"
chart.style = 13
chart.x_axis.title = 'Fecha'
chart.y_axis.title = 'Valor'
chart.y_axis.crossAx = 500
chart.x_axis = DateAxis(crossAx=100)
chart.x_axis.number_format ='yyyy/mm/dd'
chart.x_axis.majorTimeUnit = "days"
data = Reference(ws, min_col=1, min_row=1, max_row=int(cantidad))
chart.add_data(data, titles_from_data=True)
dates = Reference(ws, min_col=2, min_row=2, max_row=int(cantidad))
chart.set_categories(dates)
ws.add_chart(chart, "E2")


wb.save('/home/tomas/Escritorio/negrito/Datos2/'+sys.argv[2]+'/'+a+"f"+'.xlsx')



